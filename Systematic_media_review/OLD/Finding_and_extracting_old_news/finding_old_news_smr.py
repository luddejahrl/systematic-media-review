## Google search implementation
from pickle import NONE
import pandas as pd
import json
import requests
import csv
import openpyxl #import read_excel
import nltk
from newspaper import Article
import numpy as np

#------------------------------------Defining GCE (Google custom search engine) parametres-----------------------------------------
# get the API KEY here: https://developers.google.com/custom-search/v1/overview
## These are profile specifik and you need to get your own to do the costume search (not hard, just search google if you're unsure)
API_KEY = "AIzaSyCXVQGXI-YUWlCXcLbhZk-zv0zuxlgqBUE"

# get your Search Engine ID on your CSE control panel (from same place as the API_KEY)
SEARCH_ENGINE_ID = "f142e568df70f4f3f"

#------------------------------------/Defining GCE (Google custom search engine) parametres-----------------------------------------

file_loc = '/Users/luddejahrl/Desktop/Systematic_media_review/Finding_and_extracting_old_news/Data_initial_SMR.xlsx'

# 'article date' & 'article publisher'
cols = [1, 2]

df = pd.read_excel(file_loc, usecols=cols, index_col=NONE, engine='openpyxl')

# load excel with its path
#wrkbk = openpyxl.load_workbook("data./Users/luddejahrl/Desktop/Systematic_media_review/Finding and extracting old news/Data_initial_SMR.xlsx")
  
#sh = wrkbk.active

# for row in sh.iter_rows(min_row=1, min_col=1, max_row=12, max_col=3):
#     for cell in row:
#         print(cell.value, end=" ")

#------------------------------------Defining SEARCH -----------------------------------------

df.loc[:,1]

# empty search query
#query = '"Rwanda"' + " " + str(df.loc[1, "article_title"]) + " " + str(df.loc[1, "article_publisher"])

#print(query)

# #for row in sh.iter_rows(min_row=2, min_col=1, max_row=257, max_col=):
# for row in enumerate(df):

#     # empty search query
#     query = '"Rwanda"' + " " + str(df.loc[row, "article_title"]) + " " + str(df.loc[row, "article_publisher"])
    
#     print(query)


#     start = 1

#     #print(start)
#     url = f"https://www.googleapis.com/customsearch/v1?key={API_KEY}&cx={SEARCH_ENGINE_ID}&q={query}&start={start}"

#     ## Making the API request and using the requests' json() method to automatically parse the returned JSON data to a Python dictionary:

#     # make the API request
#     data = requests.get(url).json()

#     ## Now, this data is a dictionary containing many result tags. We are only interested in "items", which are the search results. 
#     ## By default, CSE will return ten search results. Let's iterate over them:

#     # get the result items
#     search_items = data.get("items")

#     # iterate results found per page
        
#     search_item = search_items[0]

#     try:
#         long_description = search_item["pagemap"]["metatags"][0]["og:description"]
#     except KeyError:
#         long_description = "N/A"

#     # get the page title
#     title = search_item.get("title")
#     # page snippet
#     snippet = search_item.get("snippet")
#     # alternatively, you can get the HTML snippet (bolded keywords)
#     html_snippet = search_item.get("htmlSnippet")
#     # extract the page url
#     link = search_item.get("link")
#     # print the results
            
#     driver.get(link)

#     page_soup = BeautifulSoup(driver.page_source, 'lxml') #'html.parser'?        

#     print("="*10, f"Result", "="*10)
#     print("Title:", title)

#     #full_article_text = page_soup.get_text()
#     #full_article_text = page_soup.find_all(itemprop="text")
#     #full_article_text = page_soup.find_all('p')[0].get_text()
#     #full_article_text = page_soup.find_all('p') # may work as of now


#     #for i in page_soup.get_text():
#     #    full_article_text = page_soup.select("body div p")[i].get_text()

#     #full_article_text = page_soup.select("body div p")
#     #full_article_text # list of elements
#     #full_article_text = page_soup.findAll(text=True)

#     # >>> txt = """\
#     # ... <p>Red</p>
#     # ... <p><i>Blue</i></p>
#     # ... <p>Yellow</p>
#     # ... <p>Light <b>green</b></p>
#     # ... """
#     # >>> import BeautifulSoup
#     # >>> BeautifulSoup.__version__
#     # '3.0.7a'
#     # >>> soup = BeautifulSoup.BeautifulSoup(txt)
#     # >>> for node in soup.findAll('p'):
#     # ...     print ''.join(node.findAll(text=True))

#     #txt = ""

#     # for node in page_soup.findAll('p'):
#     #     #print(''.join(node.findAll(text=True)))
#     #     full_article_text.join(node.findAll(text=True))

#     p_tags = page_soup.select("body div p")

#     #print(p_tags)

#     full_article_text = ""

#     #https://oxylabs.io/blog/news-scraping

#     for each in p_tags: 
#         full_article_text += str(each.get_text()) + "\n" # concatenate 
        
#     #FOLLOW WITH NEWSPAPER OR NEWSPAPER3K TO REMOVE COOKIES TEXT ETC?

#     # article = Article(full_article_text)
#     # article = Article(url)
#     # article.download()
#     # article.parse()
#     # nltk.download('punkt')
#     # article.nlp()

#     # article.authors
#     # article.publish_date
#     # print(article.text)
        



#     #within html -> within body (as deep as you can go) ---- [html body p] or [body p]

    
    
    










