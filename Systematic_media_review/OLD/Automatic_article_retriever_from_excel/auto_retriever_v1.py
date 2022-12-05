## Google search implementation
from pickle import NONE
from typing import Counter
import pandas as pd
import json
import requests
import csv
import openpyxl #import read_excel
import nltk
from newspaper import Article
import numpy as np

#------------------------------------Defining GCE (Google custom search engine) parametres-----------------------------------------
# get the API KEY here: https://developers.google.com/custom-search/v1/overview
## These are profile specifik and you need to get your own to do the costume search (not hard, just search google if you're unsure)
API_KEY = "AIzaSyCXVQGXI-YUWlCXcLbhZk-zv0zuxlgqBUE"

# get your Search Engine ID on your CSE control panel (from same place as the API_KEY)
SEARCH_ENGINE_ID = "f142e568df70f4f3f"

page = 1 #INDEX 0 (becomes one) for first page due to loop iteration

# HOW MANY TOTAL PAGES 
pages = 1

# HOW MANY ARTICLES PER PAGE (Maximum of 10)
articles = 5

#number article
article_nr = page

#start
start = page

#excludeTerms = "wikipedia, cdc, featureshoot"


#------------------------------------/Defining GCE (Google custom search engine) parametres-----------------------------------------

file_loc_initial = '/Users/luddejahrl/Desktop/Systematic_media_review/Automatic_article_retriever_from_excel/Data_initial_SMR.xlsx'

file_loc_retrieved = '/Users/luddejahrl/Desktop/Systematic_media_review/Automatic_article_retriever_from_excel/Data_retrieved_SMR.xlsx'

query = ""

country_string = '"Rwanda"'

# load excel with its path (initial)
wrkbk_intitial = openpyxl.load_workbook(file_loc_initial)

# load excel with its path (retrieved)
wrkbk_retrieved = openpyxl.load_workbook(file_loc_retrieved)

sh_initial = wrkbk_intitial.active

sh_retrieved = wrkbk_retrieved.active

counter_not_found = 0
counter_found = 0


#ITERATE THROUGH THE EXCEL FILE TO GET THE ARTICLES FOR THE HEADLINES
for row_initial, row_retrieved in zip(sh_initial.iter_rows(min_row=2, min_col=1, max_row=257, max_col=4), sh_retrieved.iter_rows(min_row=2, min_col=1, max_row=257, max_col=6)):
    
    #INCLUDE PUBLISHER?
    query = country_string + " " + str(row_initial[0].value) + " " + str(row_initial[1].value) + " 2010..2020"
    
    #print(query)
    #print(str(row_retrieved[2].value) + " " + str(row_retrieved[3].value))
        
    start = (page - 1) * articles + 1
    
    url = f"https://www.googleapis.com/customsearch/v1?key={API_KEY}&cx={SEARCH_ENGINE_ID}&q={query}&start={start}"

    # make the API request
    data = requests.get(url).json()
    
    # get the result items
    search_items = data.get("items")
    
    #ITERATE THROUGH THE RESULTS OF THE PAGE 
    for i, search_item in enumerate(search_items, start) or []:
        
        print("Ping")
        
        try:
            long_description = search_item["pagemap"]["metatags"][0]["og:description"]
        except KeyError:
                long_description = "N/A"
    
        # get the page title
        title = search_item.get("title")
        
        # extract the page url
        link = search_item.get("link")
        
        #-----Newspaper3k setup--------
        article = Article(link)
        article.download()
        
        try:
            article.parse() 
            article.nlp()
            
            title = article.title #REMOVE? to try googles title instead
            link = article.url
            text = article.text 
            
            #extracting date and formatting for comparrison
            date = str(article.publish_date.year) + "-" + str(article.publish_date.month) + "-" + str(article.publish_date.day)
            
            if title == row_initial[0].value and date == row_initial[3].value:
                row_retrieved[2].value = text
                row_retrieved[3].value = link
                counter_found += 1
                print("good")
        except:
            pass
        
        #NO articles were found fit the requirements
        if i == articles:
            #print("no article found")
            row_retrieved[2].value = "N/A"
            row_retrieved[3].value = "N/A"
            counter_not_found += 1
            print("bad")
            

print ("Process is DONE, number of articles correctly found: ", counter_found, "articles not found: ",  counter_not_found)
            
            
        
